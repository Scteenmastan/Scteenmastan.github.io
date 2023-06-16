import openpyxl

sheet = openpyxl.load_workbook('exam.xlsx').active

html = """<!DOCTYPE html>
<html xmlns="http://www.w3.org/1999/xhtml" xml:lang="en" lang="en">
<head>
    <title>中国近代史</title>
    <meta name='viewport' content='width=device-width, initial-scale=1' />

    <link href="https://www.oxfordlearnersdictionaries.com/external/styles/ringlinks.css?version=2.3.48" rel="stylesheet" type="text/css" />
    <link href="https://www.oxfordlearnersdictionaries.com/external/styles/interface.css?version=2.3.48" rel="stylesheet" type="text/css" />
    <link href="https://www.oxfordlearnersdictionaries.com/external/styles/jquery.lightbox-0.5.css?version=2.3.48" rel="stylesheet" type="text/css" />
    <link href="https://www.oxfordlearnersdictionaries.com/external/styles/header.css?version=2.3.48" rel="stylesheet" type="text/css" />
    <link href="https://www.oxfordlearnersdictionaries.com/external/styles/oald10.css?version=2.3.48" rel="stylesheet" type="text/css" />
    <link href="https://www.oxfordlearnersdictionaries.com/external/styles/responsive.css?version=2.3.48" rel="stylesheet" type="text/css" />
    <link href="https://fonts.googleapis.com/css?family=Open+Sans:700i" rel="stylesheet" type='text/css'>

    <link rel="stylesheet" href="index.css">  
    <link rel="stylesheet" href="magic-text.css">
    <link rel="stylesheet" href="ratio.css">
    <link rel="stylesheet" href="type.css">
    <link rel="stylesheet" href="checkbox.css">
    <script defer type="text/javascript" src="index.js"></script>

    <link rel="icon" type="image/x-icon" href="https://www.oxfordlearnersdictionaries.com/external/images/favicon.ico?version=2.3.48"/>
</head>

<body >
    <div id="ox-container">

        <div id="ox-header" class="">
    <div id="flex-header-container">
        <div id="flex-header" class="responsive_container">
            <a id="h-title" href="https://www.oxfordlearnersdictionaries.com/">
                <div class="old_logo" title="" alt="Oxford Learner's Dictionaries" style="height: 15px; vertical-align: middle;padding-right: 15px;"></div>
            </a>
            <span id="filler" class="fg main_nav"></span>
            <div id="smartphone-menu" class="menu_button responsive_display_on_smartphone">
                <span class="sr-only">Toggle navigation</span>
                <span class="icon-bar"></span>
                <span class="icon-bar"></span>
                <span class="icon-bar"></span>
            </div>
            <a id="h-redeem" target="_blank" href="https://account.oup.com/redeem" class="main_nav responsive_hide_on_smartphone flex-header-links">Redeem</a>
            <a id="h-upgrade" href="https://www.oxfordlearnersdictionaries.com/upgrade/" class="main_nav responsive_hide_on_smartphone flex-header-links">Upgrade</a>
            <a id="h-help" href="https://www.oxfordlearnersdictionaries.com/faq/" class="main_nav responsive_hide_on_smartphone flex-header-links">Help</a>
                                        <form name="loginForm" action='https://www.oxfordlearnersdictionaries.com/account/login' method="post">
                	<input type="hidden" name="callbackUrl" value="https://www.oxfordlearnersdictionaries.com/definition/english/many?q=many" />
                	<a id="h-sign-in" href="#" onclick="document.forms.loginForm.submit(); return false;" class="rounded-button main_nav responsive_hide_on_smartphone flex-header-links">
                        Sign in
                    </a>
                </form>
                    </div>
    </div>

    <hr class="hr_nav_first" style="margin: 0;">

    <div id="flex-menu" class=" responsive_container header-menu logout responsive_hide_on_smartphone">
        <ul class="flex-menu-desktop-nav">
            <li >
                <span id="menu-dictionaries" class="top-toolbar-active  link-right menu-elem">Dictionaries</span>
                <div class="menu-dropdown">
                    <ul>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/definition/">Dictionaries home</a></li>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/definition/english/">English</a></li>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/definition/american_english/">American English</a></li>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/definition/academic/">Academic</a></li>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/definition/collocations/">Collocations</a></li>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/translate/schulwoerterbuch/">German-English</a></li>
                    </ul>
                </div>
            </li>
            <li >
                <span id="menu-grammar" class=" link-right menu-elem">Grammar</span>
                <div class="menu-dropdown">
                    <ul>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/grammar/">Grammar home</a></li>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/grammar/practical-english-usage/">Practical English Usage</a></li>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/grammar/online-grammar/">Learn & Practise Grammar (Beta)</a></li>
                    </ul>
                </div>
            </li>
            <li >
                <span id="menu-wordlists" class=" link-right menu-elem">Download</span>
                <div class="menu-dropdown">
                    <ul>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/wordlists/"><code>.pdf</code> file</a></li>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/mywordlist/"><code>.md</code> file</a></li>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/topic/"><code>.html</code> file</a></li>
                    </ul>
                </div>
            </li>
            <li >
                <span id="menu-resources" class=" link-right menu-elem">About</span>
                <div class="menu-dropdown">
                    <ul>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/resources/">Help center</a></li>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/text-checker/">FAQ</a></li>
                    </ul>
                </div>
            </li>
        </ul>
    </div>
    <div id="header-menu-modal" class="header-modal">
        <div id="panel-smartphone">
                           <span><a id="h-sign-in-mobile" href="#" onclick="document.forms.loginForm.submit(); return false;" class="rounded-button">
                    Sign in
                </a></span><hr style="margin: 0;">
                        <div>
                <span class="top-toolbar-active  link-right menu-elem">Dictionaries</span>
                <div class="menu-dropdown-smartphone">
                    <ul>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/definition/">Dictionaries home</a></li>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/definition/english/">English</a></li>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/definition/american_english/">American English</a></li>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/definition/academic/">Academic</a></li>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/definition/collocations/">Collocations</a></li>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/translate/schulwoerterbuch/">German-English</a></li>
                    </ul>
                </div>
            </div>
            <hr style="margin: 0;">
            <div>
                <span class=" link-right menu-elem">Grammar</span>
                <div class="menu-dropdown-smartphone">
                    <ul>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/grammar/">Grammar home</a></li>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/grammar/practical-english-usage/">Practical English Usage</a></li>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/grammar/online-grammar/">Learn & Practise Grammar (Beta)</a></li>
                    </ul>
                </div>
            </div>
            <hr style="margin: 0;">
            <div>
                <span class=" link-right menu-elem">Word Lists</span>
                <div class="menu-dropdown-smartphone">
                    <ul>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/wordlists/">Word Lists home</a></li>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/mywordlist/">My Word Lists</a></li>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/topic/">Topics</a></li>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/wordlist/new_words">Recent additions</a></li>
                    </ul>
                </div>
            </div>
            <hr style="margin: 0;">
            <div>
                <span class=" link-right menu-elem">Resources</span>
                <div class="menu-dropdown-smartphone">
                    <ul>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/resources/">Resources home</a></li>
                        <li><a href="https://www.oxfordlearnersdictionaries.com/text-checker/">Text Checker</a></li>
                    </ul>
                </div>
            </div>
            <hr style="margin: 0;">
            <a id="h-redeem" target="_blank" href="https://account.oup.com/redeem" class="main_nav">Redeem</a>
            <a id="h-upgrade" href="https://www.oxfordlearnersdictionaries.com/upgrade/" class="main_nav">Upgrade</a>
            <a id="h-help" href="https://www.oxfordlearnersdictionaries.com/faq/" class="main_nav">Help</a>
        </div>
    </div>
    <div id="searchbar" class="responsive_container">
        <div class="mainsearch">
            <form id="search-form" method="get" action='https://www.oxfordlearnersdictionaries.com/search/english/direct/'>
                <div id="dictionarySelector">
                    <div id="select_div">
                        <div>Search</div>
                    </div>
                </div>
                <div class="searchfield">
                    <input type="text" id="q" name="q"
                        class="searchfield_input"/>
					<label for="q" style="position: absolute; width: 1px; height: 1px; margin: -1px; padding: 0; overflow: hidden; clip: rect(0, 0, 0, 0); border: 0; font-family: 'Source Sans Pro', sans-serif;">Enter search text</label>
                    <div id="keyboard_letters"></div>
                </div>
                <div class="inputSuggestions"></div>
                <label id="search-btn" class="oup_icons" title="Search">
                    <input type="button" value="" onclick="search()" />
                </label>
            </form>
        </div>
    </div>
</div>





        <div class="responsive_container  xenglish">

            <div id="topslot_container" data-iaw="container">
                        <div id='ad_topslot' class='am-default '>
    </div>
                </div>

            <div class="responsive_row">
                <div id="ad_leftslot_container" class="ad_leftslot_container responsive_entry_left">    <div class="ac_leftslot sticky" data-iaw="container">    <div id='ad_leftslot' class='am-default '>
    </div>
    </div></div>

                                <div class="responsive_entry_center">
                                    <div class="responsive_entry_center_wrap">                
                        <div id="ox-wrapper" class="responsive_entry_center_left">
                            <div id="main_column" class="responsive_row">
                                <div id="main-container" class="main-container">
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                

                                
                                
"""

id = 0
for r in range(2, 758):
    id += 1
    if sheet.cell(row=r, column=2).value == '单选题':
        html += f"""            <div class='single' id='question{id}'>
    <h1 class="headword">第 {id} 题</h1>
    <div id="ring-links-box">
        <p>{sheet.cell(row=r, column=3).value}</p>
    </div>
    <div class='options idioms bulgy-radios'>
        <label class="option A">
            <input type="radio" name="{r}"/>
            <span class="radio"></span>
            <span class="label">{sheet.cell(row=r, column=9).value}</span>
        </label><br>
        <label class="option B">
            <input type="radio" name="{r}"/>
            <span class="radio"></span>
            <span class="label">{sheet.cell(row=r, column=10).value}</span>
        </label><br>
        <label class="option C">
            <input type="radio" name="{r}"/>
            <span class="radio"></span>
            <span class="label">{sheet.cell(row=r, column=11).value}</span>
        </label><br>
        <label class="option D">
            <input type="radio" name="{r}"/>
            <span class="radio"></span>
            <span class="label">{sheet.cell(row=r, column=12).value}</span>
        </label><br>
    </div>
    <div class="collapse answer">
        <span class="unbox">
            <span class="box_title">Check</span>
            <span class="body"></span>
        </span>
        <div class='check_answer'>{sheet.cell(row=r, column=4).value}</div>
    </div>
</div><br>\n"""
    elif sheet.cell(row=r, column=2).value == '多选题':
        html += f"""            <div class='multiple' id='question{id}'>
    <h1 class="headword">第 {id} 题</h1>
    <div id="ring-links-box">
        <p>{sheet.cell(row=r, column=3).value}</p>
    </div>
    <div class='options idioms checkbox-wrapper-19'>         
        <label class="option A"><input type="checkbox" id="{r}A" name="{r}" />
        <label for="{r}A" class="check-box"></label>
        <span class="checkboxtext">{sheet.cell(row=r, column=9).value}</span></label><br>
        <label class="option B"><input type="checkbox" id="{r}B" name="{r}" />
        <label for="{r}B" class="check-box"></label>
        <span class="checkboxtext">{sheet.cell(row=r, column=10).value}</span></label><br>
        <label class="option C"><input type="checkbox" id="{r}C" name="{r}" />
        <label for="{r}C" class="check-box"></label>
        <span class="checkboxtext">{sheet.cell(row=r, column=11).value}</span></label><br>
        <label class="option D"><input type="checkbox" id="{r}D" name="{r}" />
        <label for="{r}D" class="check-box"></label>
        <span class="checkboxtext">{sheet.cell(row=r, column=12).value}</span></label><br>
    </div>
    <div class="collapse answer">
        <span class="unbox">
            <span class="box_title">Check</span>
            <span class="body"></span>
        </span>
        <div class='check_answer'>{sheet.cell(row=r, column=4).value}</div>
    </div>
</div><br>\n"""
    elif sheet.cell(row=r, column=2).value == '判断题':
        html += f"""            <div class='decide' id='question{id}'>
    <h1 class="headword">第 {id} 题</h1>
    <div id="ring-links-box">
        <p>{sheet.cell(row=r, column=3).value}</p>
    </div>
    <div class='options idioms bulgy-radios'>
        <label class="option 正确">
            <input type="radio" name="{r}"/>
            <span class="radio"></span>
            <span class="label">{sheet.cell(row=r, column=9).value}</span>
        </label><br>
        <label class="option 错误">
            <input type="radio" name="{r}"/>
            <span class="radio"></span>
            <span class="label">{sheet.cell(row=r, column=10).value}</span>
        </label><br>
    </div>
    <div class="collapse answer">
        <span class="unbox">
            <span class="box_title">Check</span>
            <span class="body"></span>
        </span>
        <div class='check_answer'>{"正确" if sheet.cell(row=r, column=4).value == "A" else "错误"}</div>
    </div>
</div><br>\n"""
    elif sheet.cell(row=r, column=2).value == '填空题':
        html += f"""            <div class='fill' id='question{id}'>
    <h1 class="headword">第 {id} 题</h1>
    <div id="ring-links-box">
        <p>{sheet.cell(row=r, column=3).value}</p>
    </div>
    <div class="collapse answer">
        <span class="unbox">
            <span class="box_title">Check</span>
            <span class="body"></span>
        </span>
        <div class='check_answer'>{sheet.cell(row=r, column=9).value}</div>
    </div>
</div><br>\n"""
    else:
        id -= 1
        
html += """        <script language="JavaScript" type="text/javascript" src="https://www.oxfordlearnersdictionaries.com/common.js?version=2.3.48"></script>
    </div>
</body>
</html>"""

with open('index.html', 'w', encoding='utf-8') as f:
    f.write(html)
